from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill, Color
from math import ceil
__author__ = 'Max'


# This function takes the hour, day and year data from an aermod line
# example input:
# *** CONCURRENT  1-HR AVERAGE CONCENTRATION   VALUES ENDING WITH HOUR  5 FOR DAY   1 OF 2016 ***
# returns list with data in form [hour, day, year]
def find_time_data_from_line(line):
    # removing blank spaces at beginning and end of line and then splitting the line into words
    line = line.strip()
    line_split = line.split(' ')

    # removing all blank items since there are some double and triple spaces in the line
    line_split = list(word for word in line_split if word != '')

    # finding the number value after the keyword 'HOUR'
    hour_index = line_split.index('HOUR')
    hour = line_split[hour_index + 1]

    # finding the number value after the keyword 'DAY'
    day_index = line_split.index('DAY')
    day = line_split[day_index + 1]

    # finding the number value after the keyword 'OF' for the year
    year_index = line_split.index('OF')
    year = line_split[year_index + 1]

    return [hour, day, year]


# this function sets up the headers at the top of the excel spreadsheet file
# inputs
# worksheet: openpyxl Workbook spread sheet *that has already been opened* (workbook.active)
# receptor information for formatting
def spreadsheet_setup_discrete(number_receptors, worksheet, receptor_x_list, receptor_y_list):
    # setting up time headers
    worksheet.cell(row=1, column=1, value='Hour')
    worksheet.cell(row=1, column=2, value='Day')
    worksheet.cell(row=1, column=3, value='Year')

    # setting up emissions data headers depending how many receptors there are
    # also sets up titles for different concentration averages
    # current_column is used to keep track of the columns so it spaces headers out correctly
    current_column = 5
    for i in range(number_receptors):
        # formatting headers and average value titles
        receptor_name = '[' + str(receptor_x_list[i]) + ', ' + str(receptor_y_list[i]) + ']'
        worksheet.cell(row=1, column=current_column - 1, value="Receptor at:")
        worksheet.cell(row=2, column=current_column - 1, value=receptor_name + " Average (excluding '0' values):")
        worksheet.cell(row=5, column=current_column - 1, value=receptor_name + " Average (including '0' values):")
        worksheet.cell(row=1, column=current_column, value=receptor_name)
        worksheet.column_dimensions[get_column_letter(current_column)].width = 10
        worksheet.column_dimensions[get_column_letter(current_column - 1)].width = 40
        current_column += 2

    # setting up informational cell
    worksheet.cell(row=9, column=4).value = "concentrations in micrograms/meters^3"
    worksheet['D9'].fill = PatternFill(fgColor=Color('FFFF00'), fill_type='solid', patternType='solid')


# this function finds all the hour times that AERMOD used to calculate concentration
# inputs
# output_file: the aermod output file name
# output_spreadsheet: openpyxl Workbook spread sheet *that has already been opened* (workbook.active)
def find_time_lines(output_file_name, output_spreadsheet):
    output_file = open(output_file_name, 'r')
    curr_row = 2
    for line in output_file:
        # looks through each line to see if it is the correct line that has time data
        # example for type of line that AERMOD produces
        #  *** CONCURRENT  1-HR AVERAGE CONCENTRATION   VALUES ENDING WITH HOUR  5 FOR DAY   1 OF 2016 ***
        line_strip = line.strip()
        line_split = line_strip.split(' ')

        # checks each line to see if it is the correct line
        if line_split[0] == '***' and line_split[1] == 'CONCURRENT':
            # find_time_data_from_line returns time data in form [hour, day, year]
            time_data = find_time_data_from_line(line)
            output_spreadsheet.cell(row=curr_row, column=1).value = int(time_data[0])
            output_spreadsheet.cell(row=curr_row, column=2).value = int(time_data[1])
            output_spreadsheet.cell(row=curr_row, column=3).value = int(time_data[2])
            curr_row += 1
    output_file.close()


# this function adds concentrations and their coordinates to the spreadsheet
# takes in single emission point data
# spreadsheet: openpyxl Workbook spread sheet *that has already been opened* (workbook.active)
def add_concentration_to_spreadsheet(emission_data, current_row, receptor_number, spreadsheet):
    # formatting is setup for excel spreadsheet with time information in first 3 columns
    #   remove the '3 + ' to just have emissions data in the final spreadsheet
    column_number = 3 + receptor_number*2
    spreadsheet.cell(row=current_row, column=column_number).value = float(emission_data)


# this function removes the last values from the spreadsheet since they are averages and not calculated data
def format_average_values_discrete(output_spreadsheet, number_receptors):
    # loops through receptor columns to find last data point
    current_column = 3
    max_row = output_spreadsheet.max_row
    for receptor in range(number_receptors):
        # finding the column the current receptors data is at, and the column to put the average value in
        current_column += 2
        average_column = current_column - 1

        # finding last element in list, setting to correct average cell
        current_average = output_spreadsheet.cell(row=max_row, column=current_column).value
        output_spreadsheet.cell(row=3, column=average_column).value = current_average

        # adding averages including zeros calculated by excel
        column_identifier = get_column_letter(current_column) + ":" + get_column_letter(current_column)
        output_spreadsheet.cell(row=6, column=average_column).value = "=AVERAGE("+column_identifier+")"

    # deleting last row
    output_spreadsheet.delete_rows(max_row)


# this function finds all concentrations at all receptors that AERMOD calculated
# inputs
# output_file_name: the aermod output file name
# number_receptors: the number of receptors specified for formatting excel spreadsheet
# output_spreadsheet: openpyxl Workbook spread sheet *that has already been opened* (workbook.active)
def find_concentration_lines_discrete(output_file_name, number_receptors, output_spreadsheet):
    output_file = open(output_file_name, 'r')
    number_data_lines_to_read = int(ceil(number_receptors / 2.0))
    current_row = 1
    # uses while loop since you cant readline() if iterating lines in a for loop in python 2.7
    while True:
        line = output_file.readline()
        if not line:
            break
        current_line_stripped = line.strip()
        current_line_split = current_line_stripped.split(' ')

        # checking to see if the line is two lines up from the data
        #   this appears to be the most viable way to collect the emissions data
        if current_line_split[0] == 'X-COORD':
            # keeps track of what row to enter data on
            current_row += 1

            # parses down to get to the correct line
            output_file.readline()

            # counts which receptor the parser is on for formatting data on excel sheet
            receptor_count = 1

            # since there are two receptor values per line, loops through lines if there are more than two receptors
            for i in range(number_data_lines_to_read):
                # formatting line to get location and emission data from each line
                data_line = output_file.readline()
                data_line = data_line.strip()
                data_line_split = data_line.split(' ')

                # getting rid of extraneous blank items in the list
                data_line_split = list(word for word in data_line_split if word != '')

                # adding the emission results data point to the spreadsheet
                add_concentration_to_spreadsheet(data_line_split[2], current_row, receptor_count, output_spreadsheet)
                receptor_count += 1

                # if there are two data points on the line, adds the second data point as well
                if len(data_line_split) == 6:
                    add_concentration_to_spreadsheet(data_line_split[5], current_row, receptor_count,
                                                     output_spreadsheet)
                    receptor_count += 1

    # deleting last row of data since it is averages computed by AERMOD
    # saves data and places near top
    # adds excel calculations that calculate concentration averages including 0
    format_average_values_discrete(output_spreadsheet, number_receptors)
    output_file.close()


# this function finds the average concentration values for a grid receptor system
# output_spreadsheet: openpyxl Workbook spread sheet *that has already been opened* (workbook.active)
def find_grid_concentration_average(output_spreadsheet, output_file_name):
    output_file = open(output_file_name, 'r')
    # variables to keep track of location in spreadsheet
    code_end = False
    current_column_header = 1
    current_column = 1
    column_start_index = 0

    # uses while loop since you cant readline() if iterating lines in a for loop in python 2.7
    while True:
        line = output_file.readline()
        # breaks at the end
        if not line:
            break
        # processing line
        line_stripped = line.strip()
        line_split = line_stripped.split(' ')
        # removing blank entries
        line_split = list(word for word in line_split if word != '')
        # checks to see if at the end of the aermod output file to get annual concentration data
        if len(line_split) > 5:
            if line_split[0] == '***' and line_split[1] == 'THE' and line_split[2] == 'ANNUAL':
                # marking that the code has parsed to the last entry
                code_end = True

        # checking if x coordinate header line
        if len(line_split) > 0 and code_end and line_split[0] == "(METERS)":
            # removing non data in line split list
            line_split = line_split[2:]

            for column in line_split:
                # adding to the column count for spreadsheet formatting
                # this for loop adds the headers of the receptor locations
                current_column_header += 1
                output_spreadsheet.cell(row=1, column=current_column_header).value = float(column)

            # parsing down two lines to get to concentration data
            output_file.readline()
            output_file.readline()
            current_row = 1

            # parsing all y - coord values (rows)
            while True:
                current_row += 1
                line = output_file.readline()
                line_stripped = line.strip()
                line_split = line_stripped.split(' ')
                # removing the '|' from data table
                del line_split[1]
                # gets rid of the y - coord value from list if there is already one there
                if column_start_index > 0:
                    del line_split[0]
                # removing blank entries
                line_split = list(word for word in line_split if word != '')
                # checks if current line contains concentration data
                try:
                    line_split = [float(num) for num in line_split]
                    # resets column if there is more data to read
                    current_column = column_start_index
                except ValueError:
                    # once a set of data has been processed, sets index to current column so data keeps
                    #   being put to the right in the correct order
                    if current_column > 0:
                        column_start_index = current_column
                    break

                for column in line_split:
                    # adding data to spreadsheet from table
                    current_column += 1
                    output_spreadsheet.cell(row=current_row, column=current_column).value = column

    output_file.close()

