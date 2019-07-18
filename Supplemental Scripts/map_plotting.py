import matplotlib.pyplot as plt
import matplotlib
from openpyxl import load_workbook
import numpy as np

# this script creates a 2d heat map of emission concentration data based on AERMOD outputs *only* using a
#   GRID RECEPTOR SYSTEM, will not work for discrete receptors
# MAKE SURE IN run_framework.py THE RECEPTOR SYSTEM IS SET TO GRID --- WILL NOT WORK OTHERWISE
# you can change the colormap of the heatmap in line 198, by changing the 'cmap=' variable

# to add to the aermod framework mainframe complete the following instructions:
# 1) add the following import statement to the very top of mainframe.py
"""
from map_plotting import create_concentration_map
"""
# 2) add the following function call statement to the very bottom of run_aermod_framework() in mainframe.py
# 2) *** You can change the names of the files to whatever you want***
"""
create_concentration_map(spreadsheet_name, "concentration heatmap")
"""

# if you want to just run this script on a set of excel data that you already have, simply add the line of
#   code above, the code from section '2)' to the very bottom of this script and change spreadsheet_name
#   to the actual name of the spreadsheet data, for example "AERMOD concentration output.xlsx"

"""
heatmap and heatmap_annotations functions are not mine
code comes from matplotlib website, I am not claiming this code is mine
Title: Creating annotated heatmaps
Availability: https://matplotlib.org/3.1.1/gallery/images_contours_and_fields/image_annotated_heatmap.html#sphx-glr-gallery-images-contours-and-fields-image-annotated-heatmap-py
"""


def heatmap(data, row_labels, col_labels, ax=None,
            cbar_kw={}, cbarlabel="", **kwargs):
    """
    Create a heatmap from a numpy array and two lists of labels.

    Parameters
    ----------
    data
        A 2D numpy array of shape (N, M).
    row_labels
        A list or array of length N with the labels for the rows.
    col_labels
        A list or array of length M with the labels for the columns.
    ax
        A `matplotlib.axes.Axes` instance to which the heatmap is plotted.  If
        not provided, use current axes or create a new one.  Optional.
    cbar_kw
        A dictionary with arguments to `matplotlib.Figure.colorbar`.  Optional.
    cbarlabel
        The label for the colorbar.  Optional.
    **kwargs
        All other arguments are forwarded to `imshow`.
    """

    if not ax:
        ax = plt.gca()

    # Plot the heatmap
    im = ax.imshow(data, **kwargs)

    # Create colorbar
    cbar = ax.figure.colorbar(im, ax=ax, **cbar_kw)
    cbar.ax.set_ylabel(cbarlabel, rotation=-90, va="bottom", fontsize=18)

    # We want to show all ticks...
    ax.set_xticks(np.arange(data.shape[1]))
    ax.set_yticks(np.arange(data.shape[0]))
    # ... and label them with the respective list entries.
    ax.set_xticklabels(col_labels)
    ax.set_yticklabels(row_labels)

    # Let the horizontal axes labeling appear on top.
    ax.tick_params(top=False, bottom=True,
                   labeltop=False, labelbottom=True)

    # Rotate the tick labels and set their alignment.
    plt.setp(ax.get_xticklabels(), rotation=-30, ha="right",
             rotation_mode="anchor")
    plt.xticks(ha='left')
    # Turn spines off and create white grid.
    for edge, spine in ax.spines.items():
        spine.set_visible(False)

    ax.set_xticks(np.arange(data.shape[1]+1)-.5, minor=True)
    ax.set_yticks(np.arange(data.shape[0]+1)-.5, minor=True)
    ax.grid(which="minor", color="w", linestyle='-', linewidth=1)
    ax.tick_params(which="minor", bottom=False, left=False)

    return im, cbar


def annotate_heatmap(im, data=None, valfmt="{x:.2f}",
                     textcolors=["black", "white"],
                     threshold=None, **textkw):
    """
    A function to annotate a heatmap.

    Parameters
    ----------
    im
        The AxesImage to be labeled.
    data
        Data used to annotate.  If None, the image's data is used.  Optional.
    valfmt
        The format of the annotations inside the heatmap.  This should either
        use the string format method, e.g. "$ {x:.2f}", or be a
        `matplotlib.ticker.Formatter`.  Optional.
    textcolors
        A list or array of two color specifications.  The first is used for
        values below a threshold, the second for those above.  Optional.
    threshold
        Value in data units according to which the colors from textcolors are
        applied.  If None (the default) uses the middle of the colormap as
        separation.  Optional.
    **kwargs
        All other arguments are forwarded to each call to `text` used to create
        the text labels.
    """

    if not isinstance(data, (list, np.ndarray)):
        data = im.get_array()

    # Normalize the threshold to the images color range.
    if threshold is not None:
        threshold = im.norm(threshold)
    else:
        threshold = im.norm(data.max())/2.

    # Set default alignment to center, but allow it to be
    # overwritten by textkw.
    kw = dict(horizontalalignment="center",
              verticalalignment="center")
    kw.update(textkw)

    # Get the formatter in case a string is supplied
    if isinstance(valfmt, str):
        valfmt = matplotlib.ticker.StrMethodFormatter(valfmt)

    # Loop over the data and create a `Text` for each "pixel".
    # Change the text's color depending on the data.
    texts = []
    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            kw.update(color=textcolors[int(im.norm(data[i, j]) > threshold)])
            text = im.axes.text(j, i, valfmt(data[i, j], None), **kw)
            texts.append(text)

    return texts


# this function creates a heatmap from an excel spreadsheet with the following assumptions
#   the spreadsheet is a complete rectangle of data
#   the top row contains x-axis headers for the x location of the data below it
#   the furthest left column contains y-axis headers for the y location of the data to the right
# the following is an example of the format that a 3x4 spreadsheet should be in
# this is the format AERMOD concentration outputs would be in, notice x axis is increasing and y axis is decreasing
# |  |-1| 0| 1| 2|
# | 1| @| @| @| @|
# | 0| @| @| @| @|
# |-1| @| @| @| @|
def create_concentration_map(spreadsheet_file_name, figure_file_name):
    # this is so the figure won't be shown, only saved as a file
    matplotlib.use('agg')
    # setup
    workbook = load_workbook(spreadsheet_file_name)
    worksheet = workbook.active
    number_rows = worksheet.max_row
    number_cols = worksheet.max_column
    number_y_data = number_rows - 1
    number_x_data = number_cols - 1
    spreadsheet_data = []
    header_list_x = []
    header_list_y = []

    # for loop puts all data into one 2d array
    # starts at 2 to *NOT* include header data
    # since openpyxl starts at 1, not 0, have to go one extra in range of data
    # when adding data, we subtract two to start the data in the '0' column
    for row in range(2, number_rows+1):
        spreadsheet_data.append([])
        for col in range(2, number_cols+1):
            spreadsheet_data[row - 2].append(worksheet.cell(row=row, column=col).value)

    # creating lists of x and y headers
    for row in range(2, number_rows+1):
        header_list_y.append(worksheet.cell(row=row,column=1).value)
    for col in range(2, number_cols+1):
        header_list_x.append(worksheet.cell(row=1,column=col).value)

    # turning data into numpy array
    spreadsheet_data_numpy = np.array(spreadsheet_data)

    # creating heatmap
    fig, ax = plt.subplots()
    im, cbar = heatmap(data=spreadsheet_data_numpy, row_labels=header_list_y, col_labels=header_list_x,
                       ax=ax, cbarlabel="Emissions Concentration (micrograms/meters$^3$)", cmap='RdBu')

    # # uncomment the following line if you want to see individual concentration data on each grid data point
    # annotate_heatmap(im,data=spreadsheet_data_numpy, valfmt="{x:.1f}")

    # title, labels, saving formatting settings
    plt.title("Concentration Heat Map", fontsize=24)
    fig.canvas.set_window_title("Concentration Heat Map")
    plt.xlabel("X Axis Receptor Locations", fontsize=18)
    plt.ylabel("Y Axis Receptor Locations", fontsize=18)
    x_size_inches = number_x_data*2/3  # can easily change if too small/big
    y_size_inches = number_y_data*2/3
    plt.gcf().set_size_inches(x_size_inches, y_size_inches)
    plt.savefig(figure_file_name+'.png')

    # # show plot if wanted, just saves as png by default
    # plt.show()

    # # print data for debugging
    # print(header_list_x)
    # for row in range(number_y_data):
    #    print(str(header_list_y[row])+str(spreadsheet_data[row]))


